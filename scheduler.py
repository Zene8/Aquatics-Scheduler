# scheduler.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime
import re

def to_dt(time):
    if isinstance(time, str):
        time = time.strip().lower()
        for fmt in ("%I%p", "%I:%M%p", "%H:%M:%S", "%H:%M"):
            try:
                return datetime.combine(datetime.today(), datetime.strptime(time, fmt).time())
            except ValueError:
                continue
        raise ValueError(f"Unsupported time format: {time}")
    return datetime.combine(datetime.today(), time)

def is_available(row, time):
    return to_dt(row['AM Start']) <= to_dt(time) <= to_dt(row['AM End'])

def parse_cell_content(content):
    student_count = 0
    instructor_names = []
    tokens = [token.strip() for token in re.split(r'[\n,/]', str(content)) if token.strip()]
    for token in tokens:
        match = re.match(r"(.+?)\s*[\\/]?\s*(\d+)$", token)
        if match:
            name = match.group(1).strip()
            count = int(match.group(2))
            student_count = count
            instructor_names.append((name, count))
        elif token.isdigit():
            student_count = int(token)
        else:
            instructor_names.append((token, None))
    return instructor_names, student_count

def apply_schedule_formatting(ws):
    """Apply custom formatting to the schedule output"""
    from openpyxl.styles import PatternFill, Border, Side
    
    # Define colors
    colors = {
        'light_blue': 'ccffff',
        'light_pink': 'ecc4f2', 
        'lime_green': '57fa06',
        'light_red': 'ff6161',
        'cyan': '29eef3',
        'dark_blue': '4d93d9',
        'purple': 'a41eb2',
        'gray': '808080',
        'yellow': 'ffff00',
        'deep_blue': '00b0f0',  # Deeper blue for filled classes
        'white': 'FFFFFF'
    }
    
    # Define black border
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Header formatting
    header_colors = {
        'Time': colors['light_blue'],
        'Starters': colors['light_blue'],
        'P1': colors['light_pink'],
        'P2': colors['light_pink'],
        'P3': colors['light_pink'],
        'Y1': colors['lime_green'],
        'Y2': colors['lime_green'],
        'Y3': colors['lime_green'],
        'PSL': colors['light_red'],
        'STRK4': colors['cyan'],
        'STRK5': colors['cyan'],
        'STRK6': colors['cyan'],
        'TN BCS AD BCS': colors['dark_blue'],
        'TN STRK AD STRK': colors['dark_blue'],
        'TN/AD BSCS': colors['dark_blue'],
        'TN/ AD STRKS': colors['dark_blue'],
        'CNDTNG': colors['purple'],
        'brk': colors['gray']
    }
    
    # Apply header colors and borders
    for col_idx, header in enumerate(ws[1], 1):
        if header.value in header_colors:
            header.fill = PatternFill(start_color=header_colors[header.value], end_color=header_colors[header.value], fill_type="solid")
        header.border = black_border
    
    # Apply cell colors and borders based on content
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # Apply black border to all cells
            cell.border = black_border
            
            if cell.value is None:
                cell.fill = PatternFill(start_color=colors['yellow'], end_color=colors['yellow'], fill_type="solid")
                continue
                
            cell_value = str(cell.value).strip()
            
            # Determine cell color based on content and column
            cell_color = colors['yellow']  # default
            
            # Check if it's in Time column or brk column (no highlighting)
            col_header = ws.cell(row=1, column=cell.column).value
            if col_header and (col_header == 'Time' or col_header.lower() == 'brk'):
                cell_color = colors['white']  # No highlighting for Time and brk columns
            elif cell_value == "":
                cell_color = colors['yellow']  # Yellow for unfilled/empty classes
            elif cell_value and cell_value != "":
                # Check if it's in PSL column
                if col_header and 'PSL' in str(col_header).upper():
                    cell_color = colors['light_red']
                else:
                    cell_color = colors['deep_blue']  # Deeper blue for filled classes
            
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

def generate_am_schedule(template_df, availability_df, template_path=None, manual_mode=False):
    if hasattr(template_path, 'read'):
        template_bytes = template_path.read()
        template_path = BytesIO(template_bytes)

    wb = load_workbook(filename=template_path)
    ws = wb.active

    # Fix the header issue by ensuring proper column names
    headers = [cell.value for cell in ws[1]]
    
    # If the first column doesn't have a proper name, set it to "Time"
    if not headers[0] or headers[0] == "Unnamed: 0":
        headers[0] = "Time"
        ws.cell(row=1, column=1, value="Time")
    
    time_slots = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]

    instructor_assignments = {name: [] for name in availability_df['Name']}
    instructor_roles = availability_df.set_index("Name")['Role'].to_dict()
    assignments_by_time = {str(t): [] for t in time_slots}
    total_breaks = {name: 0 for name in availability_df['Name']}
    last_class_by_instructor = {name: None for name in availability_df['Name']}

    # TEMPLATE CONSISTENCY: Create mapping of template instructors to classes
    template_mapping = {}
    template_instructors_set = set()  # Track all instructors from template
    if not manual_mode:
        # Build template mapping from the uploaded template
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers))):
            time_cell = row[0]
            time_str = str(time_cell.value)
            
            for j, cell in enumerate(row[1:], start=1):
                class_name = headers[j]
                if not class_name or class_name.lower() == "brk":
                    continue
                    
                content = str(cell.value) if cell.value else ""
                if not content.strip():
                    continue
                    
                parsed_entries, student_count = parse_cell_content(content)
                template_instructors = [name for name, _ in parsed_entries if name]
                
                if template_instructors:
                    template_mapping[f"{time_str}_{class_name}"] = {
                        'instructors': template_instructors,
                        'student_count': student_count
                    }
                    # Add to set of all template instructors
                    for inst in template_instructors:
                        template_instructors_set.add(inst.lower())

    # Create availability mapping for efficient lookup
    available_instructors = {name.lower(): name for name in availability_df['Name']}
    template_available = {inst: available_instructors.get(inst.lower()) for inst in template_instructors_set if inst.lower() in available_instructors}
    non_template_instructors = [name for name in availability_df['Name'] if name.lower() not in template_instructors_set]

    # First pass: assign at least one instructor to each class
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers))):
        time_cell = row[0]
        time_str = str(time_cell.value)

        for j, cell in enumerate(row[1:], start=1):
            class_name = headers[j]
            if not class_name:
                continue
            class_name = class_name.strip()
            if class_name.lower() == "brk":
                continue

            content = str(cell.value) if cell.value else ""
            
            # For manual mode, only process cells that have content (selected classes)
            if manual_mode:
                if not content.strip():
                    # Skip empty cells in manual mode - they should remain blank
                    continue
                else:
                    # This is a selected class that needs an instructor assigned
                    assigned = []
                    shadows = []
                    
                    def can_be_assigned(inst_name, inst_row):
                        if not is_available(inst_row, time_cell.value):
                            return False
                        if class_name in inst_row.get("Can't Teach", []):
                            return False
                        if inst_name in assignments_by_time[time_str]:
                            return False
                        return True
                    
                    # Try to assign a regular instructor
                    available_pool = availability_df.copy()
                    available_pool['Assignments'] = available_pool['Name'].map(lambda name: len(instructor_assignments[name]))
                    available_pool = available_pool.sort_values(by='Assignments')
                    
                    for _, inst_row in available_pool.iterrows():
                        inst_name = inst_row['Name']
                        if not can_be_assigned(inst_name, inst_row):
                            continue
                        if inst_row['Role'] == 'Shadow':
                            continue
                        assigned.append(inst_name)
                        last_class_by_instructor[inst_name] = class_name
                        assignments_by_time[time_str].append(inst_name)
                        instructor_assignments[inst_name].append(time_str)
                        break
                    
                    # Assign a shadow if possible
                    if assigned:
                        for _, inst_row in available_pool.iterrows():
                            inst_name = inst_row['Name']
                            if inst_name in assigned:
                                continue
                            if not can_be_assigned(inst_name, inst_row):
                                continue
                            if inst_row['Role'] == 'Shadow':
                                shadows.append(inst_name)
                                assignments_by_time[time_str].append(inst_name)
                                instructor_assignments[inst_name].append(time_str)
                                break
                    
                    result_lines = []
                    for a in assigned:
                        result_lines.append(a)
                    for s in shadows:
                        result_lines.append(s)
                    
                    formatted_text = "\n".join(result_lines) if result_lines else "unfilled"
                    cell.value = formatted_text
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    continue
            
            # TEMPLATE MODE: Optimized assignment with priority logic
            if not manual_mode:
                # Only process cells that have content in the template (existing classes)
                if not content.strip():
                    continue  # Skip empty cells - don't fill them
                
                assigned = []
                shadows = []
                
                def can_be_assigned(inst_name, inst_row):
                    if not is_available(inst_row, time_cell.value):
                        return False
                    if class_name in inst_row.get("Can't Teach", []):
                        return False
                    if inst_name in assignments_by_time[time_str]:
                        return False
                    return True
                
                # Get template data for this specific class
                template_key = f"{time_str}_{class_name}"
                student_count = 0
                template_instructors = []
                
                if template_key in template_mapping:
                    template_data = template_mapping[template_key]
                    template_instructors = template_data['instructors']
                    student_count = template_data['student_count']
                
                # STEP 1: Try to assign template instructors first (if available)
                for template_inst in template_instructors:
                    # Check if this template instructor is available
                    if template_inst.lower() in available_instructors:
                        inst_name = available_instructors[template_inst.lower()]
                        inst_row = availability_df[availability_df['Name'] == inst_name].iloc[0]
                        
                        if can_be_assigned(inst_name, inst_row):
                            if inst_row['Role'] == 'Shadow':
                                if assigned:  # Only add shadow if we have a regular instructor
                                    shadows.append(inst_name)
                                    assignments_by_time[time_str].append(inst_name)
                                    instructor_assignments[inst_name].append(time_str)
                            else:
                                assigned.append(inst_name)
                                last_class_by_instructor[inst_name] = class_name
                                assignments_by_time[time_str].append(inst_name)
                                instructor_assignments[inst_name].append(time_str)
                
                # STEP 2: Fill remaining slots with NON-TEMPLATE instructors first (priority)
                needed = len(template_instructors) if template_instructors else 1
                while len(assigned) < needed:
                    filled = False
                    
                    # Try non-template instructors first
                    for inst_name in non_template_instructors:
                        if inst_name in assigned:
                            continue
                        inst_row = availability_df[availability_df['Name'] == inst_name].iloc[0]
                        if not can_be_assigned(inst_name, inst_row):
                            continue
                        if inst_row['Role'] == 'Shadow':
                            continue
                        assigned.append(inst_name)
                        last_class_by_instructor[inst_name] = class_name
                        assignments_by_time[time_str].append(inst_name)
                        instructor_assignments[inst_name].append(time_str)
                        filled = True
                        break
                    
                    # If no non-template instructors available, try template instructors
                    if not filled:
                        available_pool = availability_df.copy()
                        available_pool['Assignments'] = available_pool['Name'].map(lambda name: len(instructor_assignments[name]))
                        available_pool = available_pool.sort_values(by='Assignments')
                        
                        for _, inst_row in available_pool.iterrows():
                            inst_name = inst_row['Name']
                            if inst_name in assigned:
                                continue
                            if not can_be_assigned(inst_name, inst_row):
                                continue
                            if inst_row['Role'] == 'Shadow':
                                continue
                            assigned.append(inst_name)
                            last_class_by_instructor[inst_name] = class_name
                            assignments_by_time[time_str].append(inst_name)
                            instructor_assignments[inst_name].append(time_str)
                            filled = True
                            break
                    
                    if not filled:
                        break
                
                # STEP 3: Add shadows if possible (prioritize non-template instructors)
                for inst_name in non_template_instructors:
                    if inst_name in assigned or inst_name in shadows:
                        continue
                    inst_row = availability_df[availability_df['Name'] == inst_name].iloc[0]
                    if not can_be_assigned(inst_name, inst_row):
                        continue
                    if inst_row['Role'] == 'Shadow' and assigned:
                        shadows.append(inst_name)
                        assignments_by_time[time_str].append(inst_name)
                        instructor_assignments[inst_name].append(time_str)
                        break
                
                # If no non-template shadows, try template instructors
                if not shadows and assigned:
                    for _, inst_row in availability_df.iterrows():
                        inst_name = inst_row['Name']
                        if inst_name in assigned or inst_name in shadows:
                            continue
                        if not can_be_assigned(inst_name, inst_row):
                            continue
                        if inst_row['Role'] == 'Shadow' and assigned:
                            shadows.append(inst_name)
                            assignments_by_time[time_str].append(inst_name)
                            instructor_assignments[inst_name].append(time_str)
                            break
                
                # Format output
                result_lines = []
                for a in assigned:
                    line = f"{a} ({student_count})" if student_count > 0 else f"{a}"
                    result_lines.append(line)
                for s in shadows:
                    result_lines.append(s)
                
                formatted_text = "\n".join(result_lines) if result_lines else "unfilled"
                cell.value = formatted_text
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                continue

    # Double up logic and breaks only if not manual mode
    if not manual_mode:
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers))):
            time_str = str(row[0].value)
            for j, cell in enumerate(row[1:], start=1):
                class_name = headers[j]
                if not class_name:
                    continue
                class_name = class_name.strip()
                if class_name.lower() == "brk":
                    continue
                if not cell.value or cell.value == "unfilled":
                    continue

                current_names = [line.split(" (")[0] for line in str(cell.value).split("\n")]
                parsed_entries, student_count = parse_cell_content(cell.value)
                is_psl = "PSL" in class_name.upper()
                if is_psl or student_count < 5:
                    continue

                if len(current_names) < 2:
                    available_pool = availability_df.copy()
                    available_pool = available_pool.sort_values(by='Name')
                    for _, inst_row in available_pool.iterrows():
                        inst_name = inst_row['Name']
                        if inst_name in current_names:
                            continue
                        if not is_available(inst_row, time_str):
                            continue
                        if inst_row['Role'] == 'Shadow':
                            continue
                        if inst_name in assignments_by_time[time_str]:
                            continue
                        current_names.append(inst_name)
                        assignments_by_time[time_str].append(inst_name)
                        instructor_assignments[inst_name].append(time_str)
                        break

                    updated = "\n".join(current_names)
                    cell.value = updated
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if any(h for h in headers if h and h.lower() == "brk"):
            brk_col = [h.lower() if h else "" for h in headers].index("brk") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                time_str = str(row[0].value)
                not_teaching = []
                for inst_name in instructor_assignments:
                    inst_row = availability_df[availability_df['Name'] == inst_name].iloc[0]
                    if time_str not in instructor_assignments[inst_name] and is_available(inst_row, time_str):
                        not_teaching.append(inst_name)

                required_breaks = [name for name in not_teaching if total_breaks[name] < 1 and len(instructor_assignments[name]) >= 4]
                for name in required_breaks:
                    total_breaks[name] += 1

                row[brk_col - 1].value = ", ".join(not_teaching)
                row[brk_col - 1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Handle breaks for manual mode
    if manual_mode and any(h for h in headers if h and h.lower() == "brk"):
        brk_col = [h.lower() if h else "" for h in headers].index("brk") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            time_str = str(row[0].value)
            not_teaching = []
            for inst_name in instructor_assignments:
                inst_row = availability_df[availability_df['Name'] == inst_name].iloc[0]
                if time_str not in instructor_assignments[inst_name] and is_available(inst_row, time_str):
                    not_teaching.append(inst_name)

            row[brk_col - 1].value = ", ".join(not_teaching)
            row[brk_col - 1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply custom formatting to the schedule
    apply_schedule_formatting(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    preview_data = []
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
        preview_data.append(row)

    return output, preview_data
