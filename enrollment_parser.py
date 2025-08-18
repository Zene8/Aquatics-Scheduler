# enrollment_parser.py - Parse enrollment Excel files to generate schedule templates

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os

class EnrollmentParser:
    def __init__(self):
        self.class_mapping = {
            "Swim Starters Stage A Exploration": "Starters",
            "Swim Starters Stage B Exploration": "Starters",
            "Preschool Stage 1 Water Acclimation": "P1",
            "Preschool Stage 2 Water Movement": "P2", 
            "Preschool Stage 3 Water Stamina": "P3",
            "School Age Stage 1 Water Acclimation": "Y1",
            "School Age Stage 2 Water Movement": "Y2",
            "School Age Stage 3 Water Stamina": "Y3",
            "School Age Stage 4 Stroke Introduction": "STRK4",
            "School Age Stage 5 Stroke Development": "STRK5", 
            "School Age Stage 6 Stroke Mechanics": "STRK6",
            "Teen / Adult Swim Basics": "TN/AD BSCS",
            "Teen / Adult Swim Strokes": "TN/AD STRKS",
            "Aquatic Conditioning Swim Team Prep": "CNDTNG"
        }
        
        self.am_times = ["8:35", "9:10", "9:45", "10:20", "11:00", "11:35", "12:10"]
        self.pm_times = ["4:10", "4:45", "5:20", "5:55", "6:30", "7:05"]
    
    def parse_group_lessons(self, file_path, target_date, session_mode):
        """Parse group lessons Excel file and extract classes for the target date"""
        print("=" * 80)
        print("🚨 ENROLLMENT PARSER CALLED 🚨")
        print(f"File: {file_path}")
        print(f"Target Date: {target_date}")
        print(f"Target Date Type: {type(target_date)}")
        print(f"Session Mode: {session_mode}")
        print("=" * 80)
        
        try:
            df = pd.read_excel(file_path)
            
            # Convert target_date to date object
            if isinstance(target_date, str):
                target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                target_date = target_date.date()
            
            target_day = target_date.strftime('%A')
            print(f"🎯 PROCESSING: Date {target_date} = {target_day}")
            print(f"🎯 LOOKING FOR: {target_day} classes in {session_mode} mode")
            
            classes = []
            
            print(f"Processing {len(df)} rows for {target_day} {session_mode} session")
            
            for idx, row in df.iterrows():
                # STEP 1: Check if this row has the target day in "Day of Week" column
                day_of_week = str(row.get('Day of Week', '')).strip()
                if not day_of_week or day_of_week == 'nan':
                    continue
                
                # Handle semicolon-separated days (e.g., "Tuesday; Thursday")
                # Split by semicolon first, then by comma, and clean up each day
                days_list = []
                for part in day_of_week.split(';'):
                    for day in part.split(','):
                        days_list.append(day.strip())
                
                # Check if target day is in the list (exact match)
                if target_day not in days_list:
                    continue
                
                print(f"Row {idx}: Found {target_day} in '{day_of_week}' (parsed as {days_list})")
                
                # STEP 2: Check if "Status" column says "Approved"
                status = str(row.get('Status', '')).strip()
                if status != "Approved":
                    print(f"Row {idx}: Status '{status}' is not 'Approved' - SKIPPING")
                    continue
                
                print(f"Row {idx}: Status is 'Approved' ✓")
                
                # STEP 3: Check "Enrollment Status" - eliminate "Under Minimum Enrollment"
                enrollment_status = str(row.get('Enrollment Status', '')).strip()
                if enrollment_status == "Under Minimum Enrollment":
                    print(f"Row {idx}: Enrollment Status '{enrollment_status}' - SKIPPING (not enough students)")
                    continue
                
                print(f"Row {idx}: Enrollment Status '{enrollment_status}' ✓")
                
                # STEP 4: Check "Course Start Time" for AM/PM session
                start_time_str = str(row.get('Course Start Time', '')).strip()
                if not start_time_str or start_time_str == 'nan':
                    print(f"Row {idx}: No start time - SKIPPING")
                    continue
                
                # Check if time matches the session mode (AM/PM)
                start_time_str_upper = start_time_str.upper()
                if session_mode == "AM" and "PM" in start_time_str_upper:
                    print(f"Row {idx}: Time '{start_time_str}' is PM but we want AM - SKIPPING")
                    continue
                elif session_mode == "PM" and "AM" in start_time_str_upper:
                    print(f"Row {idx}: Time '{start_time_str}' is AM but we want PM - SKIPPING")
                    continue
                
                print(f"Row {idx}: Time '{start_time_str}' matches {session_mode} session ✓")
                
                # STEP 5: Get course name and map to class type
                course_name = str(row.get('Course Option: Course Option Name', '')).strip()
                if course_name not in self.class_mapping:
                    print(f"Row {idx}: Course '{course_name}' not in mapping - SKIPPING")
                    continue
                
                class_type = self.class_mapping[course_name]
                print(f"Row {idx}: Course '{course_name}' maps to '{class_type}' ✓")
                
                # STEP 6: Get "Total Enrolled" count
                enrollment = int(row.get('Total Enrolled', 0))
                
                # STEP 7: Parse the time to get hour:minute format
                try:
                    # Extract time part (e.g., "9:10" from "9:10 AM")
                    time_part = start_time_str.split()[0]
                    hour_str, minute_str = time_part.split(':')
                    hour = int(hour_str.strip())
                    minute = int(minute_str.strip())
                    
                    # Convert to 24-hour format if needed
                    if "PM" in start_time_str_upper and hour != 12:
                        hour += 12
                    elif "AM" in start_time_str_upper and hour == 12:
                        hour = 0
                    
                    formatted_time = f"{hour:02d}:{minute:02d}"
                    
                    print(f"Row {idx}: ✓ FINAL RESULT - {class_type} at {formatted_time} with {enrollment} students")
                    
                    classes.append({
                        'class_type': class_type,
                        'start_time': formatted_time,
                        'enrollment': enrollment
                    })
                    
                except (ValueError, IndexError) as e:
                    print(f"Row {idx}: Error parsing time '{start_time_str}': {e} - SKIPPING")
                    continue
            
            print(f"Total classes found after all filters: {len(classes)}")
            print("=" * 80)
            return classes
            
        except Exception as e:
            print(f"Error parsing group lessons: {e}")
            return []
    
    def parse_private_lessons(self, file_path, target_date, session_mode):
        """Parse private lessons Excel file and extract classes for the target date"""
        try:
            df = pd.read_excel(file_path)
            
            # Convert target_date to date object
            if isinstance(target_date, str):
                target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                target_date = target_date.date()
            
            private_lessons = []
            
            print(f"Looking for private lessons on {target_date} in {session_mode} mode")
            
            for _, row in df.iterrows():
                course_option = str(row.get('Course Option', '')).strip()
                start_date_str = str(row.get('Start Date', '')).strip()
                start_time_str = str(row.get('Start Time', '')).strip()
                
                if not start_date_str or start_date_str == 'nan':
                    continue
                
                try:
                    start_date = pd.to_datetime(start_date_str).date()
                    
                    # Check if this lesson should be included
                    include_lesson = False
                    
                    if course_option == "Private Swim Lessons":
                        course_end_date = start_date + timedelta(weeks=4)
                        if start_date <= target_date <= course_end_date:
                            include_lesson = True
                    elif course_option == "Swim Lessons - Private Package (4 pack)":
                        days_diff = (target_date - start_date).days
                        if days_diff in [0, 7, 14, 21]:
                            include_lesson = True
                    
                    if include_lesson and start_time_str and start_time_str != 'nan':
                        # Parse time - handle formats like "9:10 AM" or "4:10 PM"
                        start_time_str = start_time_str.upper()
                        
                        if ':' in start_time_str:
                            time_part = start_time_str.split()[0]  # Get "9:10" from "9:10 AM"
                            hour_str, minute_str = time_part.split(':')
                            
                            hour = int(hour_str.strip())
                            minute = int(minute_str.strip())
                            
                            # Convert to 24-hour format based on AM/PM
                            if 'PM' in start_time_str and hour != 12:
                                hour += 12
                            elif 'AM' in start_time_str and hour == 12:
                                hour = 0
                            
                            # Check session time range
                            if session_mode == "AM" and 8 <= hour <= 12:
                                first_name = str(row.get('First Name', '')).strip()
                                age = str(row.get('Age', '')).strip()
                                
                                print(f"Found PSL for {first_name} ({age}) at {hour:02d}:{minute:02d}")
                                private_lessons.append({
                                    'first_name': first_name,
                                    'age': age,
                                    'start_time': f"{hour:02d}:{minute:02d}"
                                })
                            elif session_mode == "PM" and (hour >= 15 or hour < 8):
                                first_name = str(row.get('First Name', '')).strip()
                                age = str(row.get('Age', '')).strip()
                                
                                print(f"Found PSL for {first_name} ({age}) at {hour:02d}:{minute:02d}")
                                private_lessons.append({
                                    'first_name': first_name,
                                    'age': age,
                                    'start_time': f"{hour:02d}:{minute:02d}"
                                })
                        
                except (ValueError, TypeError) as e:
                    print(f"Error parsing PSL data: {e}")
                    continue
            
            print(f"Total private lessons found: {len(private_lessons)}")
            return private_lessons
            
        except Exception as e:
            print(f"Error parsing private lessons: {e}")
            return []
    
    def generate_schedule_template(self, group_classes, private_lessons, session_mode):
        """Generate schedule template from parsed enrollment data"""
        try:
            template_path = "assets/Blank_MTWT_Template1.xlsx"
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            wb = load_workbook(template_path)
            ws = wb.active
            
            time_slots = self.am_times if session_mode == "AM" else self.pm_times
            
            # Clear existing content (except headers)
            for row in range(2, ws.max_row + 1):
                for col in range(2, ws.max_column + 1):
                    ws.cell(row=row, column=col, value="")
            
            # Update time column
            for i, time_slot in enumerate(time_slots):
                if i + 2 <= ws.max_row:
                    ws.cell(row=i + 2, column=1, value=time_slot)
            
            # Get headers to map class types to columns
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            header_to_col = {header: col for col, header in enumerate(headers, 1)}
            
            # Process group classes - place them at their exact times
            for class_info in group_classes:
                class_type = class_info['class_type']
                start_time = class_info['start_time']
                enrollment = class_info['enrollment']
                
                # Find the closest time slot for this class
                closest_time = self._find_closest_time_slot(start_time, time_slots)
                if closest_time and class_type in header_to_col:
                    time_row = time_slots.index(closest_time) + 2
                    col = header_to_col[class_type]
                    ws.cell(row=time_row, column=col, value=enrollment)
            
            # Process private lessons - place them at their exact times
            if private_lessons and 'PSL' in header_to_col:
                psl_col = header_to_col['PSL']
                
                # Group PSLs by their closest time slots
                psl_by_time = {}
                for lesson in private_lessons:
                    start_time = lesson.get('start_time', '')
                    if start_time:
                        closest_time = self._find_closest_time_slot(start_time, time_slots)
                        if closest_time:
                            if closest_time not in psl_by_time:
                                psl_by_time[closest_time] = []
                            psl_by_time[closest_time].append(f"{lesson['first_name']} ({lesson['age']})")
                
                # Place PSLs at their times
                for time_slot in time_slots:
                    if time_slot in psl_by_time:
                        time_row = time_slots.index(time_slot) + 2
                        if time_row <= ws.max_row:
                            ws.cell(row=time_row, column=psl_col, value='; '.join(psl_by_time[time_slot]))
            
            # Save the template
            wb.save(template_path)
            return True
            
        except Exception as e:
            print(f"Error generating schedule template: {e}")
            return False
    
    def get_template_preview(self, group_classes, private_lessons, session_mode, target_date=None):
        """Generate a preview of the schedule template"""
        try:
            time_slots = self.am_times if session_mode == "AM" else self.pm_times
            
            # Define all columns in the correct order
            all_columns = [
                'Time', 'Starters', 'P1', 'P2', 'P3', 'Y1', 'Y2', 'Y3', 'PSL',
                'STRK4', 'STRK5', 'STRK6', 'TN/AD BSCS', 'TN/AD STRKS', 'CNDTNG', 'brk'
            ]
            
            # Create DataFrame
            preview_data = {col: [''] * len(time_slots) for col in all_columns}
            preview_data['Time'] = time_slots
            
            # Fill in group classes - place them at their exact times
            for class_info in group_classes:
                class_type = class_info['class_type']
                start_time = class_info['start_time']
                enrollment = class_info['enrollment']
                
                # Find the closest time slot for this class
                closest_time = self._find_closest_time_slot(start_time, time_slots)
                if closest_time and class_type in preview_data:
                    time_idx = time_slots.index(closest_time)
                    preview_data[class_type][time_idx] = enrollment
            
            # Fill in private lessons - place them at their exact times
            if private_lessons:
                psl_by_time = {}
                for lesson in private_lessons:
                    start_time = lesson.get('start_time', '')
                    if start_time:
                        closest_time = self._find_closest_time_slot(start_time, time_slots)
                        if closest_time:
                            if closest_time not in psl_by_time:
                                psl_by_time[closest_time] = []
                            psl_by_time[closest_time].append(f"{lesson['first_name']} ({lesson['age']})")
                
                # Fill PSL column based on exact times
                for time_slot in time_slots:
                    if time_slot in psl_by_time:
                        time_idx = time_slots.index(time_slot)
                        preview_data['PSL'][time_idx] = '; '.join(psl_by_time[time_slot])
            
            df = pd.DataFrame(preview_data)
            
            # Add a title row with date and session info
            if target_date:
                # Convert target_date to date object if it's a string
                if isinstance(target_date, str):
                    target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
                elif isinstance(target_date, datetime):
                    target_date = target_date.date()
                
                # Format the date nicely
                formatted_date = target_date.strftime('%A, %B %d, %Y')
                title = f"Generated Schedule Preview for {formatted_date} {session_mode} Session"
                
                # Add the title as a caption attribute to the DataFrame
                df.attrs['title'] = title
            
            return df
            
        except Exception as e:
            print(f"Error generating preview: {e}")
            return pd.DataFrame()
    
    def _find_closest_time_slot(self, start_time, time_slots):
        """Find the closest time slot for a given start time"""
        try:
            hour, minute = map(int, start_time.split(':'))
            start_minutes = hour * 60 + minute
            
            closest_time = None
            min_diff = float('inf')
            
            for time_slot in time_slots:
                slot_hour, slot_minute = map(int, time_slot.split(':'))
                slot_minutes = slot_hour * 60 + slot_minute
                diff = abs(start_minutes - slot_minutes)
                
                if diff < min_diff:
                    min_diff = diff
                    closest_time = time_slot
            
            return closest_time
        except (ValueError, TypeError):
            return None 